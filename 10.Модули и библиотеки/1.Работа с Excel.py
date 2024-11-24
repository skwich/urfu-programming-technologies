import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side


def get_vacancies(filename: str):
    return pd.read_csv("vacancies.csv", names=["name", "s_from", "s_to", "s_cur", "city", "date"])


def apply_style(cell, is_header=False):
    border = Border(left=Side("thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    cell.border = border
    if is_header:
        cell.font = Font(bold=True)


def create_statistics_of_years(wb: Workbook, vacs: pd.DataFrame):
    wb.create_sheet("Статистика по годам", 0)
    sheet = wb[wb.sheetnames[0]]
    sheet.column_dimensions["A"].width = 5
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 20

    vacs["year"] = vacs["date"].apply(lambda x: x[:4])
    vacs["s_avg"] = vacs[["s_from", "s_to"]].mean(1)
    df = vacs.groupby("year", as_index=False).agg({"s_avg": "mean", "name": "count"})
    df["s_avg"] = df["s_avg"].round().astype(int)

    data = [["Год", "Средняя зарплата", "Количество вакансий"]] + df.values.tolist()
    for i, row in enumerate(data, 1):
        year, salary, count = row
        sheet[f"A{i}"] = year
        sheet[f"B{i}"] = salary
        sheet[f"C{i}"] = count
        apply_style(sheet[f"A{i}"], i == 1)
        apply_style(sheet[f"B{i}"], i == 1)
        apply_style(sheet[f"C{i}"], i == 1)


def create_statistics_of_city(wb: Workbook, vacs: pd.DataFrame):
    sheet = wb[wb.sheetnames[1]]
    sheet.title = "Статистика по городам"
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 2
    sheet.column_dimensions["D"].width = 20
    sheet.column_dimensions["E"].width = 20

    vacs["s_avg"] = vacs[["s_from", "s_to"]].mean(1)
    stat = vacs.groupby("city", as_index=False)[["city", "name", "s_avg"]].agg(
        {"city": "first", "name": "count", "s_avg": "mean"}
    )
    stat["s_avg"] = stat["s_avg"].round().astype(int)
    stat["name"] = (stat["name"] / len(vacs) * 100).round(2)

    salaries = (
        (stat[stat["name"] >= 1].sort_values("s_avg", ascending=False))[["city", "s_avg"]].values[:10].tolist()
    )
    for i, row in enumerate([["Город", "Уровень зарплат"]] + salaries, 1):
        city, salary = row
        sheet[f"A{i}"] = city
        sheet[f"B{i}"] = salary
        apply_style(sheet[f"A{i}"], i == 1)
        apply_style(sheet[f"B{i}"], i == 1)

    shares = (
        (stat[stat["name"] >= 1].sort_values(["name", "s_avg"], ascending=[False, True]))[["city", "name"]]
        .values[:10]
        .tolist()
    )
    for i, row in enumerate([["Город", "Доля вакансий, %"]] + shares, 1):
        city, share = row
        sheet[f"D{i}"] = city
        sheet[f"E{i}"] = share
        apply_style(sheet[f"D{i}"], i == 1)
        apply_style(sheet[f"E{i}"], i == 1)


def create_report():
    csv = "vacancies.csv"
    vacs = get_vacancies(csv)
    wb = Workbook()
    create_statistics_of_years(wb, vacs)
    create_statistics_of_city(wb, vacs)
    wb.save("student_works/report.xlsx")
    wb.close()


create_report()
